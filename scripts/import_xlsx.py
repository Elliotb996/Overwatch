#!/usr/bin/env python3
"""
OVERWATCH — Excel → Supabase Import
Parses AMC_Origin_Tracker_vXX.xlsx and upserts into amc_flights table.

Usage:
    pip install openpyxl supabase python-dotenv
    python import_xlsx.py --file AMC_Origin_Tracker_v15.xlsx [--dry-run]

Run from project root. Requires SUPABASE_URL and SUPABASE_SERVICE_KEY
in environment or .env file.
"""

import argparse
import os
import sys
from datetime import datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(".env.local")

SUPABASE_URL = os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    sys.exit("Missing SUPABASE_URL / SUPABASE_SERVICE_KEY in environment")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# Column header → field name mapping
COL_MAP = {
    "Base":           "base",
    "Dep Date (UTC)": "dep_date",
    "Callsign":       "callsign",
    "ICAO Hex":       "hex",
    "Serial":         "serial",
    "Mission Code":   "mission_code",
    "First Hop":      "first_hop",
    "Via":            "via",
    "Destination":    "destination",
    "Dest Arr Date":  "dest_arr_date",
    "Dest Dep Date":  "dest_dep_date",
    "Onward":         "onward",
    "Return MC":      "return_mc",
    "Status":         "status",
    "Notes":          "notes",
}

def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s in ("—", "-", "None"): return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except: pass
    return None

def clean(val):
    if val is None: return None
    s = str(val).strip()
    return None if s in ("—", "-", "None", "") else s

def normalise_status(val):
    s = clean(val) or "ACTIVE"
    s = s.upper()
    if "COMPLETE" in s: return "COMPLETE"
    if "ACTIVE"   in s: return "ACTIVE"
    if "PENDING"  in s: return "PENDING"
    return "ACTIVE"

def parse_sheet(ws, source_label):
    rows = list(ws.iter_rows(values_only=True))
    hdr_row = None
    for i, row in enumerate(rows[:6]):
        if row[0] == "Base":
            hdr_row = i; break
    if hdr_row is None:
        print(f"  [{source_label}] No header row found — skipping")
        return []

    headers = [str(c).strip() if c else "" for c in rows[hdr_row]]
    col_idx = {COL_MAP[h]: headers.index(h) for h in COL_MAP if h in headers}
    
    records = []
    for row in rows[hdr_row + 1:]:
        base = clean(row[col_idx.get("base", 0)])
        if not base or len(base) != 4 or base.startswith("  "): continue

        mc = clean(row[col_idx["mission_code"]]) if "mission_code" in col_idx else None
        if not mc: continue

        record = {
            "base":          base.upper(),
            "dep_date":      parse_date(row[col_idx["dep_date"]]) if "dep_date" in col_idx else None,
            "callsign":      clean(row[col_idx.get("callsign", 99)]),
            "hex":           clean(row[col_idx.get("hex", 99)]),
            "serial":        clean(row[col_idx.get("serial", 99)]),
            "mission_code":  mc,
            "first_hop":     clean(row[col_idx.get("first_hop", 99)]),
            "via":           clean(row[col_idx.get("via", 99)]),
            "destination":   clean(row[col_idx.get("destination", 99)]),
            "dest_arr_date": parse_date(row[col_idx.get("dest_arr_date", 99)] if "dest_arr_date" in col_idx else None),
            "return_mc":     clean(row[col_idx.get("return_mc", 99)]),
            "status":        normalise_status(row[col_idx.get("status", 99)] if "status" in col_idx else "COMPLETE"),
            "notes":         clean(row[col_idx.get("notes", 99)]),
            "source":        "xlsx_import",
        }
        # Drop None values except required fields
        records.append({k: v for k, v in record.items() if v is not None or k in ("base", "mission_code")})

    return records


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file",    required=True, help="Path to .xlsx file")
    parser.add_argument("--dry-run", action="store_true", help="Parse without writing to Supabase")
    args = parser.parse_args()

    print(f"Opening {args.file}...")
    wb = openpyxl.load_workbook(args.file, data_only=True)
    
    all_records = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        records = parse_sheet(ws, sname)
        print(f"  [{sname}] parsed {len(records)} flights")
        all_records.extend(records)

    print(f"\nTotal: {len(all_records)} records")

    if args.dry_run:
        print("DRY RUN — not writing to Supabase.")
        for r in all_records[:5]:
            print(" ", r)
        return

    # Batch upsert (100 at a time)
    inserted = 0; updated = 0; errors = 0
    batch_size = 100
    for i in range(0, len(all_records), batch_size):
        batch = all_records[i:i+batch_size]
        try:
            result = supabase.table("amc_flights").upsert(
                batch, on_conflict="callsign,dep_date,mission_code"
            ).execute()
            inserted += len(result.data)
            print(f"  Batch {i//batch_size + 1}: {len(result.data)} upserted")
        except Exception as e:
            errors += len(batch)
            print(f"  Batch {i//batch_size + 1} ERROR: {e}")

    print(f"\nDone. {inserted} records upserted, {errors} errors.")
    if errors:
        print("Check Supabase logs for constraint violations.")


if __name__ == "__main__":
    main()
